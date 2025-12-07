import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def center_column(ws, column, horizontal='center', vertical='center', 
                 wrap_text=False, start_row=1, end_row=None):
    """
    Центрирует указанный столбец
    
    Args:
        ws: рабочий лист
        column: столбец (буква или номер)
        horizontal: горизонтальное выравнивание
        vertical: вертикальное выравнивание
        wrap_text: перенос текста
        start_row: начальная строка
        end_row: конечная строка (None = до конца)
    """
    # Преобразуем номер столбца в букву, если нужно
    if isinstance(column, int):
        col_letter = get_column_letter(column)
    else:
        col_letter = column
    
    # Определяем последнюю строку
    if end_row is None:
        end_row = ws.max_row
    
    # Создаем выравнивание
    alignment = Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=wrap_text
    )
    
    # Применяем ко всем ячейкам в столбце
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=getColumnLetterIndex(col_letter))
        cell.alignment = alignment

def getColumnLetterIndex(col_letter):
    """Преобразует букву столбца в номер"""
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(col_letter)

# ['№','Наименование','Адрес','Бит','Функция','Тип данных','Значение для записи', 'Примечание']
# Создаем данные
data = [
    ['№','Наименование','Адрес','Бит','Функция','Тип данных','Значение для записи', 'Примечание'],
    [1, '1. Состояние', '', '', '', '', '', ''],
    [2, '1.1 Статусы', '', '', '', '', '', ''],
    [3,'ЭПУ на резерве',12,2,3,'int16','',''],
    [4,'ЭПУ на инверторе',12,3,3,'int16','',''],
    [5, '1.2 Аварии', '', '', '', '', '', ''],
    [6,'Авария инвертора',13,0,3,'int16','',''],
    [7,'Выход не в норме',13,1,3,'int16','',''],
    [8, 'Измерения', '', '', '', '', '', ''],
    [9,'Напряжение на инверторе',1122,'',3,'float32','',''],
    [10,'Ток инвертора',1124,'',3,'float32','',''],
]
# data = [
#     ['ОТДЕЛ РАЗРАБОТКИ    ', '', ''],  # Строка-подзаголовок
#     ['Имя', 'Должность', 'Зарплата'],
#     ['Анна', 'Разработчик', 50000],
#     ['Борис', 'Тестировщик', 45000],
#     ['', '', ''],  # Пустая строка-разделитель
#     ['ОТДЕЛ ПРОДАЖ', '', ''],  # Еще один подзаголовок
#     ['Имя', 'Должность', 'Зарплата'],
#     ['Виктор', 'Менеджер', 60000],
#     ['Дарья', 'Аналитик', 55000],
#     ['', '', ''],  # Еще разделитель
#     ['ОТДЕЛ МАРКЕТИНГА', '', ''],
#     ['Имя', 'Должность', 'Зарплата'],
#     ['Елена', 'Дизайнер', 48000],
#     ['Федор', 'Копирайтер', 42000]
# ]

df = pd.DataFrame(data)

# Сначала сохраняем
df.to_excel('авто_подзаголовки.xlsx', index=False, header=False)

# Автоматически находим и форматируем подзаголовки
wb = load_workbook('авто_подзаголовки.xlsx')
ws = wb.active

# Установка фикс ширины столбцов
# ['№','Наименование','Адрес','Бит','Функция','Тип данных','Значение для записи', 'Примечание'],
ws.column_dimensions['A'].width = 4 
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 50

# Центрируем разные столбцы по-разному
center_column(ws, 'A', horizontal='center')
center_column(ws, 'B', horizontal='left', vertical='center', wrap_text=True, start_row=2) 
center_column(ws, 'B', horizontal='center', vertical='center', wrap_text=True, end_row=1)
center_column(ws, 'C', horizontal='center', vertical='center', wrap_text=True)
center_column(ws, 'D', horizontal='center', vertical='center', wrap_text=True)
center_column(ws, 'E', horizontal='center', vertical='center', wrap_text=True)
center_column(ws, 'F', horizontal='center', vertical='center', wrap_text=True)
center_column(ws, 'G', horizontal='center', vertical='center', wrap_text=True)
center_column(ws, 'H', horizontal='left', vertical='center', wrap_text=True, start_row=2)
center_column(ws, 'H', horizontal='center', vertical='center', wrap_text=True, end_row=1)

# Определяем критерии для подзаголовков
subheader_keywords = ['Состояние', 'Статусы', 'Аварии', 'Измерения']

for row in range(1, ws.max_row + 1):
    cell_value = ws.cell(row=row, column=2).value
    print(cell_value)
    if cell_value and any(keyword in str(cell_value) for keyword in subheader_keywords):
        # Объединяем все ячейки в строке
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ws.max_column)
        
        # Форматируем
        cell = ws.cell(row=row, column=2)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True, size=14)
        
        # Чередуем цвета для разных отделов
        # if 'РАЗРАБОТКИ' in cell_value:
        #     cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
        # elif 'ПРОДАЖ' in cell_value:
        #     cell.fill = PatternFill(start_color="ED7D31", fill_type="solid")
        # elif 'МАРКЕТИНГА' in cell_value:
        #     cell.fill = PatternFill(start_color="A9D08E", fill_type="solid")
        # else:
        # cell.fill = PatternFill(start_color="7030A0", fill_type="solid")
        
        # Форматируем заголовки столбцов (следующую строку)
        # for col in range(1, ws.max_column + 1):
        #     header_cell = ws.cell(row=row + 1, column=col)
        #     header_cell.font = Font(bold=True)
        #     header_cell.fill = PatternFill(start_color="E2EFDA", fill_type="solid")


# Определяем границы
for ws in wb.worksheets:
    # Определяем максимальные используемые строки и столбцы
    max_row = ws.max_row
    max_column = ws.max_column
    
    # Если есть данные
    if max_row > 0 and max_column > 0:
        # Создаем стиль границы
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # Применяем границы ко всем ячейкам в диапазоне
        for row in ws.iter_rows(min_row=1, max_row=max_row, 
                               min_col=1, max_col=max_column):
            for cell in row:
                cell.border = border



# Сохраняем
wb.save('авто_подзаголовки.xlsx')
print("Файл сохранен: авто_подзаголовки.xlsx")