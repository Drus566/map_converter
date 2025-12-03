# excel_exporter.py
# (Предполагается, что установлена библиотека openpyxl: pip install openpyxl)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def export_to_excel(data_lines, output_filename="output.xlsx"):
    """
    Создает Excel-файл и форматирует его.
    """
    if not data_lines:
        print("⚠️ Нет данных для экспорта в Excel.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Обработанные данные"
    
    # 1. Настройка стиля заголовка
    header_font = Font(bold=True, color="FFFFFF") # Белый, жирный
    header_fill = PatternFill(start_color="0070C0", fill_type="solid") # Синий фон

    # 2. Создание заголовка
    ws['A1'] = "ID"
    ws['B1'] = "Данные"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['B1'].font = header_font
    ws['B1'].fill = header_fill
    
    # 3. Заполнение данными
    for index, line in enumerate(data_lines, start=2): # Начинаем со 2-й строки
        ws[f'A{index}'] = index - 1 # ID
        ws[f'B{index}'] = line     # Сама строка данных

    # 4. Сохранение
    wb.save(output_filename)
    print(f"✅ Данные успешно экспортированы в Excel: {output_filename}")